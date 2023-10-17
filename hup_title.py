import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime


# 指定目标网页的URL
url = 'https://bbs.hupu.com/all-gambia'  # 将此处替换为你要爬取的网页URL

# 发送HTTP请求获取网页内容
response = requests.get(url)

# 检查响应状态码，确保请求成功
if response.status_code == 200:
    # 使用Beautiful Soup解析网页内容
    soup = BeautifulSoup(response.text, 'html.parser')

    # 获取当前日期作为表格名称
    current_date = datetime.now().strftime('%Y-%m-%d')

    # 创建一个新的Excel工作簿
    try:
        workbook = load_workbook('output.xlsx')
    except FileNotFoundError:
        workbook = Workbook()

    # 创建一个新的工作表，以当前日期命名
    sheet = workbook.create_sheet(title=current_date)

    # 设置表头
    sheet['A1'] = 'Text'
    sheet['B1'] = 'Href'

    #设置列宽
    sheet.column_dimensions['A'].width = 80
    sheet.column_dimensions['B'].width = 50

    # 查找所有具有class属性为"list-item-wrap"的<div>标签
    divs_with_class_1 = soup.find_all('div', class_='list-item-wrap')

    # 初始化行数
    row = 2

    # 遍历这些<div>标签
    for div_tag in divs_with_class_1:
        # 在每个<div>标签下查找具有class属性为"t-title"的<span>元素
        span_tag = div_tag.find('span', class_='t-title')

        # 检查是否找到了<span>元素
        if span_tag:
            # 提取<span>元素的文本内容
            span_text = span_tag.text

            # 查找<div>标签下的<a>标签
            a_tag = div_tag.find('a')

            # 检查是否找到了<a>标签
            if a_tag:
                # 提取<a>标签的href属性值
                href_attr = 'https://bbs.hupu.com/all-gambia' + a_tag.get('href')

                # 将提取的信息写入Excel文件
                sheet.cell(row=row, column=1, value=span_text)
                sheet.cell(row=row, column=2, value=href_attr)

                # 增加行数
                row += 1

    # 保存Excel文件
    workbook.save('/root/test/output.xlsx')

    print(f'信息已写入工作表：{current_date}')
else:
    print('请求失败，状态码：', response.status_code)
